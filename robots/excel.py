from openpyxl import Workbook
from openpyxl.styles import Alignment

from robots.models import Robot

from datetime import datetime, timedelta


def generate_excel_report():
    # Создаем новую книгу Excel
    wb = Workbook()

    # Определяем начало и конец периода недели
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)

    # Заголовок таблицы
    header = ['Модель', 'Версия', 'Количество за неделю']

    # Создаем страницу для каждой модели робота
    robot_models = Robot.objects.values_list('model', flat=True).distinct()
    for model in robot_models:
        ws = wb.create_sheet(model)  # Создаем страницу с именем модели

        # Устанавливаем заголовок таблицы
        ws.append(header)

        # Фильтр для роботов, созданных в течение последней недели
        robots_created_last_week = Robot.objects.filter(model=model, created__range=[start_date, end_date])

        # Создаем словарь для подсчета количества роботов по версиям
        version_count = {}

        # Подсчитаем количество роботов за неделю для каждой версии
        for robot in robots_created_last_week:
            version = robot.version
            if version in version_count:
                version_count[version] += 1
            else:
                version_count[version] = 1

        # Заполняем таблицу данными
        for version, count in version_count.items():
            ws.append([model, version, count])

        # Выравнивание текста по центру для всех ячеек
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # Удаляем стандартную страницу "Sheet"
    del wb['Sheet']

    # Сохраняем файл
    wb.save('robot_report.xlsx')
